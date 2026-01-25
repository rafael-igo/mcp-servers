#!/usr/bin/env python3
"""
Excel Server MCP
================
Leitura avançada de arquivos Excel com suporte a fórmulas e múltiplas abas.
"""

import sys
import json
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.utils import get_column_letter
from mcp.server.fastmcp import FastMCP

mcp = FastMCP("excel-server")


@mcp.tool()
def read_excel_tabs(file_path: str) -> str:
    """
    Lê todas as abas de um arquivo Excel e retorna os dados.

    Args:
        file_path: Caminho completo para o arquivo Excel

    Returns:
        JSON string com dados de todas as abas
    """
    try:
        path = Path(file_path)
        if not path.exists():
            return json.dumps({
                "success": False,
                "error": f"Arquivo não encontrado: {file_path}"
            })

        workbook = openpyxl.load_workbook(file_path, data_only=False)
        result = {
            "success": True,
            "file": str(path.name),
            "sheets": []
        }

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Extrair dados da aba
            data = []
            for row in sheet.iter_rows(values_only=True):
                # Converter valores para tipos serializáveis
                row_data = [
                    str(cell) if cell is not None else None
                    for cell in row
                ]
                data.append(row_data)

            result["sheets"].append({
                "name": sheet_name,
                "rows": len(data),
                "cols": len(data[0]) if data else 0,
                "data": data[:100]  # Limitar a 100 linhas para não sobrecarregar
            })

        workbook.close()
        return json.dumps(result, indent=2)

    except Exception as e:
        return json.dumps({
            "success": False,
            "error": str(e)
        })


@mcp.tool()
def read_excel_with_formulas(file_path: str, sheet_name: str = None) -> str:
    """
    Lê arquivo Excel preservando as fórmulas.

    Args:
        file_path: Caminho completo para o arquivo Excel
        sheet_name: Nome da aba (opcional, usa primeira se não especificado)

    Returns:
        JSON string com dados e fórmulas
    """
    try:
        path = Path(file_path)
        if not path.exists():
            return json.dumps({
                "success": False,
                "error": f"Arquivo não encontrado: {file_path}"
            })

        workbook = openpyxl.load_workbook(file_path, data_only=False)

        # Selecionar aba
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                return json.dumps({
                    "success": False,
                    "error": f"Aba '{sheet_name}' não encontrada"
                })
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        result = {
            "success": True,
            "file": str(path.name),
            "sheet": sheet.title,
            "cells": []
        }

        # Iterar sobre células com fórmulas
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_data = {
                        "address": f"{get_column_letter(cell.column)}{cell.row}",
                        "value": str(cell.value),
                        "data_type": str(cell.data_type)
                    }

                    # Incluir fórmula se existir
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        cell_data["formula"] = cell.value

                    result["cells"].append(cell_data)

        workbook.close()
        return json.dumps(result, indent=2)

    except Exception as e:
        return json.dumps({
            "success": False,
            "error": str(e)
        })


@mcp.tool()
def get_excel_metadata(file_path: str) -> str:
    """
    Retorna metadados do arquivo Excel.

    Args:
        file_path: Caminho completo para o arquivo Excel

    Returns:
        JSON string com metadados
    """
    try:
        path = Path(file_path)
        if not path.exists():
            return json.dumps({
                "success": False,
                "error": f"Arquivo não encontrado: {file_path}"
            })

        workbook = openpyxl.load_workbook(file_path, data_only=False)

        result = {
            "success": True,
            "file": str(path.name),
            "size_bytes": path.stat().st_size,
            "sheets_count": len(workbook.sheetnames),
            "sheets": []
        }

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Calcular dimensões
            max_row = sheet.max_row
            max_col = sheet.max_column

            result["sheets"].append({
                "name": sheet_name,
                "max_row": max_row,
                "max_column": max_col,
                "dimensions": f"{get_column_letter(1)}1:{get_column_letter(max_col)}{max_row}"
            })

        workbook.close()
        return json.dumps(result, indent=2)

    except Exception as e:
        return json.dumps({
            "success": False,
            "error": str(e)
        })


def main():
    """Entry point for the MCP server."""
    mcp.run(transport="stdio")


if __name__ == "__main__":
    main()
